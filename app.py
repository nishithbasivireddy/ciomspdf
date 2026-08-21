import io
import json
import hashlib
from pathlib import Path

import pandas as pd
import streamlit as st

from src.field_schema import ALL_FIELDS, CHECKBOX_SCHEMA, FIELD_SCHEMA
from src.components.cioms_expected_form.component import (
    blank_expected_values,
    normalize_expected_values,
    render_cioms_expected_form,
    validate_expected_values,
)
from src.extractors.plain_python_extractor import extract_plain_python
from src.extractors.ocr_python_extractor import extract_ocr_python
from src.extractors.custom_ml_extractor import extract_custom_ml
from src.extractors.llm_extractor import extract_llm
from src.evaluation.compare_outputs import compare_all_outputs
from src.utils.json_utils import load_json, save_json

st.set_page_config(
    page_title="PDF Form Extraction Quality Comparison",
    layout="wide"
)

st.title("PDF Form Extraction Quality Comparison")

st.warning(
    "Demo app for synthetic/sample PDFs only. Do not upload confidential, personal, medical, client, or production data."
)

st.write(
    "Upload a filled PDF form, run extraction using different approaches, compare outputs, and export the quality report."
)

Path("data").mkdir(parents=True, exist_ok=True)
Path("outputs/json").mkdir(parents=True, exist_ok=True)
Path("outputs/reports").mkdir(parents=True, exist_ok=True)
Path("outputs/images").mkdir(parents=True, exist_ok=True)


if "expected_values" not in st.session_state:
    st.session_state["expected_values"] = (
        blank_expected_values()
    )

if "expected_source" not in st.session_state:
    st.session_state["expected_source"] = (
        "Not provided"
    )



def clear_previous_outputs():
    for folder in ["outputs/json", "outputs/reports", "outputs/images"]:
        for file_path in Path(folder).glob("*"):
            if file_path.is_file():
                file_path.unlink(missing_ok=True)


def clean_display_output(data):
    hidden_keys = {
        "llm_status",
        "llm_model",
        "custom_ml_status",
        "ocr_status",
        "ocr_raw_text_available",
        "pdf_text_available",
        "raw_llm_response",
        "tesseract_path",
    }

    return {
        key: value
        for key, value in data.items()
        if key not in hidden_keys
    }


def run_with_progress(label, function_call):
    progress_bar = st.progress(0)
    status_box = st.empty()

    status_box.info(f"{label} started...")
    progress_bar.progress(20)

    with st.spinner(f"{label} in progress..."):
        progress_bar.progress(60)
        result = function_call()
        progress_bar.progress(90)

    progress_bar.progress(100)
    status_box.success(f"{label} completed.")

    return result


def run_plain_python():
    return extract_plain_python(
        pdf_path="data/filled_form.pdf",
        output_path="outputs/json/extracted_plain_python.json",
        validation_path="outputs/json/plain_python_reference_validation.json",
    )


def run_ocr_python():
    try:
        return extract_ocr_python(
            pdf_path="data/filled_form.pdf",
            output_path="outputs/json/extracted_ocr_python.json"
        )
    except Exception as error:
        fail_output = {field: "" for field in ALL_FIELDS}
        fail_output["ocr_status"] = f"Failed: {str(error)}"
        save_json(fail_output, "outputs/json/extracted_ocr_python.json")
        return fail_output


def run_custom_ml():
    return extract_custom_ml(
        pdf_path="data/filled_form.pdf",
        output_path="outputs/json/extracted_custom_ml.json"
    )


def run_llm():
    return extract_llm(
        pdf_path="data/filled_form.pdf",
        output_path="outputs/json/extracted_llm.json"
    )


CHECKBOX_FIELD_LABELS = {
    "reaction_patient_died": (
        "7. Seriousness - Patient Died"
    ),
    "reaction_hospitalisation": (
        "7. Seriousness - Hospitalisation"
    ),
    "reaction_disability": (
        "7. Seriousness - Disability or Incapacity"
    ),
    "reaction_life_threatening": (
        "7. Seriousness - Life Threatening"
    ),
    "reaction_abated_yes": (
        "20. Reaction Abated - Yes"
    ),
    "reaction_abated_no": (
        "20. Reaction Abated - No"
    ),
    "reaction_abated_na": (
        "20. Reaction Abated - Not Applicable"
    ),
    "reaction_reappeared_yes": (
        "21. Reaction Reappeared - Yes"
    ),
    "reaction_reappeared_no": (
        "21. Reaction Reappeared - No"
    ),
    "reaction_reappeared_na": (
        "21. Reaction Reappeared - Not Applicable"
    ),
    "report_source_study": (
        "24d. Report Source - Study"
    ),
    "report_source_literature": (
        "24d. Report Source - Literature"
    ),
    "report_type_initial": (
        "25a. Report Type - Initial"
    ),
    "report_type_followup": (
        "25a. Report Type - Follow-up"
    ),
    "report_source_health_professional": (
        "24d. Report Source - Health Professional"
    ),
}


def get_field_display_label(
    field: str,
) -> str:
    if field in FIELD_SCHEMA:
        return FIELD_SCHEMA[field].get(
            "label",
            field,
        )

    return CHECKBOX_FIELD_LABELS.get(
        field,
        field.replace("_", " ").title(),
    )


def build_field_level_dataframe():
    expected_path = Path(
        "outputs/json/expected_values.json"
    )

    if not expected_path.is_file():
        return pd.DataFrame()

    expected = load_json(expected_path)

    evaluation_files = {
        "Plain Python": (
            "outputs/json/"
            "plain_python_evaluation.json"
        ),
        "OCR + Python": (
            "outputs/json/"
            "ocr_plus_python_evaluation.json"
        ),
        "Custom ML + Python": (
            "outputs/json/"
            "custom_ml_plus_python_evaluation.json"
        ),
        "LLM Schema Mapping": (
            "outputs/json/"
            "llm_schema_mapping_evaluation.json"
        ),
    }

    results_by_approach = {}

    for approach, evaluation_path in (
        evaluation_files.items()
    ):
        if not Path(evaluation_path).is_file():
            results_by_approach[approach] = {}
            continue

        evaluation = load_json(
            evaluation_path
        )

        results_by_approach[approach] = {
            item.get("field", ""): item
            for item in evaluation.get(
                "field_results",
                [],
            )
        }

    rows = []

    for field in ALL_FIELDS:
        row = {
            "Field Key": field,
            "Field": get_field_display_label(
                field
            ),
            "Expected": expected.get(
                field,
                "",
            ),
        }

        for approach in evaluation_files:
            item = results_by_approach[
                approach
            ].get(
                field,
                {},
            )

            row[approach] = item.get(
                "actual",
                "",
            )

            row[
                f"{approach} Status"
            ] = item.get(
                "status",
                "Not Run",
            )

        rows.append(row)

    return pd.DataFrame(rows)


def build_confusion_dataframe():
    evaluation_files = {
        "Plain Python": (
            "outputs/json/"
            "plain_python_evaluation.json"
        ),
        "OCR + Python": (
            "outputs/json/"
            "ocr_plus_python_evaluation.json"
        ),
        "Custom ML + Python": (
            "outputs/json/"
            "custom_ml_plus_python_evaluation.json"
        ),
        "LLM Schema Mapping": (
            "outputs/json/"
            "llm_schema_mapping_evaluation.json"
        ),
    }

    rows = []

    for approach, evaluation_path in (
        evaluation_files.items()
    ):
        if not Path(evaluation_path).is_file():
            continue

        evaluation = load_json(
            evaluation_path
        )

        metrics = evaluation.get(
            "metrics",
            {},
        )

        rows.append(
            {
                "Approach": approach,
                "TP": metrics.get(
                    "true_positive",
                    0,
                ),
                "FP": metrics.get(
                    "false_positive",
                    0,
                ),
                "FN": metrics.get(
                    "false_negative",
                    0,
                ),
                "TN": metrics.get(
                    "true_negative",
                    0,
                ),
                "Exact Matches": metrics.get(
                    "matched_fields",
                    0,
                ),
                "Incorrect": metrics.get(
                    "incorrect_values",
                    0,
                ),
                "Missing": metrics.get(
                    "missing_values",
                    0,
                ),
                "Unexpected": metrics.get(
                    "unexpected_values",
                    0,
                ),
            }
        )

    return pd.DataFrame(rows)


def style_field_level_report(
    dataframe: pd.DataFrame,
):
    approach_columns = [
        "Plain Python",
        "OCR + Python",
        "Custom ML + Python",
        "LLM Schema Mapping",
    ]

    status_columns = {
        approach: f"{approach} Status"
        for approach in approach_columns
    }

    visible_columns = [
        "Field",
        "Expected",
        *approach_columns,
    ]

    display_dataframe = dataframe[
        visible_columns
    ].copy()

    def style_row(row):
        styles = [
            ""
            for _ in display_dataframe.columns
        ]

        original_row = dataframe.loc[
            row.name
        ]

        for column_index, column_name in enumerate(
            display_dataframe.columns
        ):
            if column_name == "Field":
                styles[column_index] = (
                    "font-weight: 600; "
                    "background-color: #f3f4f6;"
                )

            elif column_name == "Expected":
                styles[column_index] = (
                    "font-weight: 600; "
                    "background-color: #e5e7eb; "
                    "color: #111827;"
                )

            elif column_name in approach_columns:
                status = str(
                    original_row.get(
                        status_columns[column_name],
                        "",
                    )
                )

                if status in {
                    "Correct",
                    "Correct Blank",
                }:
                    styles[column_index] = (
                        "background-color: #dcfce7; "
                        "color: #166534;"
                    )

                elif status in {
                    "Incorrect",
                    "Missing",
                    "Unexpected",
                }:
                    styles[column_index] = (
                        "background-color: #fee2e2; "
                        "color: #991b1b;"
                    )

                else:
                    styles[column_index] = (
                        "background-color: #fef3c7; "
                        "color: #92400e;"
                    )

        return styles

    return (
        display_dataframe.style
        .apply(
            style_row,
            axis=1,
        )
        .set_properties(
            **{
                "white-space": "pre-wrap",
                "vertical-align": "top",
            }
        )
    )


def create_excel_bytes(
    summary_df,
    field_df,
    confusion_df,
):
    from openpyxl.styles import (
        Alignment,
        Font,
        PatternFill,
    )
    from openpyxl.utils import (
        get_column_letter,
    )

    output = io.BytesIO()

    approach_columns = [
        "Plain Python",
        "OCR + Python",
        "Custom ML + Python",
        "LLM Schema Mapping",
    ]

    status_columns = {
        approach: f"{approach} Status"
        for approach in approach_columns
    }

    visible_field_columns = [
        "Field",
        "Expected",
        *approach_columns,
    ]

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        summary_df.to_excel(
            writer,
            sheet_name=(
                "Approach Comparison Summary"
            ),
            index=False,
        )

        field_df[
            visible_field_columns
        ].to_excel(
            writer,
            sheet_name="Field Level Report",
            index=False,
        )

        workbook = writer.book

        summary_sheet = workbook[
            "Approach Comparison Summary"
        ]

        field_sheet = workbook[
            "Field Level Report"
        ]

        percentage_headers = {
            "Field Agreement",
            "Precision",
            "Recall",
            "F1 Score",
        }

        summary_headers = {
            cell.value: cell.column
            for cell in summary_sheet[1]
        }

        for header in percentage_headers:
            column_index = summary_headers.get(
                header
            )

            if column_index is None:
                continue

            for row_index in range(
                2,
                summary_sheet.max_row + 1,
            ):
                summary_sheet.cell(
                    row=row_index,
                    column=column_index,
                ).number_format = "0.00%"

        summary_sheet.row_dimensions[1].height = 42
        field_sheet.row_dimensions[1].height = 38

        for row_index in range(
            2,
            field_sheet.max_row + 1,
        ):
            values = [
                field_sheet.cell(
                    row=row_index,
                    column=column_index,
                ).value
                for column_index in range(
                    1,
                    field_sheet.max_column + 1,
                )
            ]

            maximum_line_count = max(
                (
                    str(value).count("\n") + 1
                    if value not in {None, ""}
                    else 1
                )
                for value in values
            )

            field_sheet.row_dimensions[
                row_index
            ].height = max(
                24,
                min(
                    90,
                    18 * maximum_line_count,
                ),
            )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F2937",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        expected_fill = PatternFill(
            fill_type="solid",
            fgColor="E5E7EB",
        )

        field_fill = PatternFill(
            fill_type="solid",
            fgColor="F3F4F6",
        )

        correct_fill = PatternFill(
            fill_type="solid",
            fgColor="DCFCE7",
        )

        incorrect_fill = PatternFill(
            fill_type="solid",
            fgColor="FEE2E2",
        )

        pending_fill = PatternFill(
            fill_type="solid",
            fgColor="FEF3C7",
        )

        for worksheet in [
            summary_sheet,
            field_sheet,
        ]:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        for row_index in range(
            2,
            field_sheet.max_row + 1,
        ):
            field_sheet.cell(
                row=row_index,
                column=1,
            ).fill = field_fill

            field_sheet.cell(
                row=row_index,
                column=1,
            ).font = Font(
                bold=True
            )

            field_sheet.cell(
                row=row_index,
                column=2,
            ).fill = expected_fill

            field_sheet.cell(
                row=row_index,
                column=2,
            ).font = Font(
                bold=True
            )

            source_row = field_df.iloc[
                row_index - 2
            ]

            for approach_index, approach in (
                enumerate(
                    approach_columns,
                    start=3,
                )
            ):
                status = str(
                    source_row.get(
                        status_columns[approach],
                        "Not Run",
                    )
                )

                target_cell = field_sheet.cell(
                    row=row_index,
                    column=approach_index,
                )

                if status in {
                    "Correct",
                    "Correct Blank",
                }:
                    target_cell.fill = (
                        correct_fill
                    )

                elif status in {
                    "Incorrect",
                    "Missing",
                    "Unexpected",
                }:
                    target_cell.fill = (
                        incorrect_fill
                    )

                else:
                    target_cell.fill = (
                        pending_fill
                    )

            for column_index in range(
                1,
                field_sheet.max_column + 1,
            ):
                field_sheet.cell(
                    row=row_index,
                    column=column_index,
                ).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        summary_widths = {
            1: 24,
            2: 18,
        }

        for column_index in range(
            3,
            summary_sheet.max_column + 1,
        ):
            summary_widths[
                column_index
            ] = 18

        for column_index, width in (
            summary_widths.items()
        ):
            summary_sheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = width

        field_widths = {
            1: 34,
            2: 28,
            3: 28,
            4: 28,
            5: 28,
            6: 28,
        }

        for column_index, width in (
            field_widths.items()
        ):
            field_sheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = width

        field_sheet.sheet_view.showGridLines = (
            False
        )

        summary_sheet.sheet_view.showGridLines = (
            False
        )

    output.seek(0)
    return output


def create_pdf_bytes(
    summary_df,
    field_df,
    confusion_df,
):
    from matplotlib.backends.backend_pdf import (
        PdfPages,
    )
    import matplotlib.pyplot as plt
    import textwrap

    output = io.BytesIO()

    approach_columns = [
        "Plain Python",
        "OCR + Python",
        "Custom ML + Python",
        "LLM Schema Mapping",
    ]

    status_columns = {
        approach: f"{approach} Status"
        for approach in approach_columns
    }

    visible_field_columns = [
        "Field",
        "Expected",
        *approach_columns,
    ]

    colors = {
        "header": "#1f2937",
        "header_text": "#ffffff",
        "field": "#f3f4f6",
        "expected": "#e5e7eb",
        "correct": "#dcfce7",
        "incorrect": "#fee2e2",
        "pending": "#fef3c7",
    }

    def display_text(
        value,
        width=24,
    ):
        if pd.isna(value):
            return ""

        text = str(value)

        return "\n".join(
            textwrap.wrap(
                text,
                width=width,
                replace_whitespace=False,
                drop_whitespace=False,
            )
        )

    def percentage_text(value):
        if pd.isna(value) or value == "":
            return ""

        return f"{float(value) * 100:.2f}%"

    def style_header(
        table,
        column_count,
    ):
        for column_index in range(
            column_count
        ):
            cell = table[
                0,
                column_index,
            ]

            cell.set_facecolor(
                colors["header"]
            )

            cell.get_text().set_color(
                colors["header_text"]
            )

            cell.get_text().set_weight(
                "bold"
            )

            cell.get_text().set_wrap(
                True
            )

    def add_summary_page(pdf):
        figure = plt.figure(
            figsize=(16.5, 11.7)
        )

        figure.subplots_adjust(
            left=0.035,
            right=0.965,
            top=0.94,
            bottom=0.05,
            hspace=0.38,
        )

        figure.suptitle(
            "CIOMS Extraction Quality Report",
            fontsize=20,
            fontweight="bold",
            y=0.975,
        )

        figure.text(
            0.5,
            0.94,
            (
                "Expected Values Source: "
                f"{st.session_state['expected_source']}"
            ),
            ha="center",
            fontsize=10,
        )

        figure.text(
            0.5,
            0.918,
            (
                "Evaluation Method: Normalized "
                "Exact-Match Field-Level Comparison"
            ),
            ha="center",
            fontsize=10,
        )

        metrics_axis = figure.add_subplot(
            2,
            1,
            1,
        )

        outcomes_axis = figure.add_subplot(
            2,
            1,
            2,
        )

        metrics_axis.axis("off")
        outcomes_axis.axis("off")

        metrics_axis.set_title(
            "Performance Metrics",
            fontsize=14,
            fontweight="bold",
            pad=12,
        )

        outcomes_axis.set_title(
            "Field Outcome Counts",
            fontsize=14,
            fontweight="bold",
            pad=12,
        )

        metrics_columns = [
            "Approach",
            "Status",
            "Field Agreement",
            "Precision",
            "Recall",
            "F1 Score",
        ]

        metrics_display = summary_df[
            [
                column
                for column in metrics_columns
                if column in summary_df.columns
            ]
        ].copy()

        for column in [
            "Field Agreement",
            "Precision",
            "Recall",
            "F1 Score",
        ]:
            if column in metrics_display.columns:
                metrics_display[column] = (
                    metrics_display[column]
                    .apply(percentage_text)
                )

        metrics_labels = [
            "Approach",
            "Status",
            "Field\nAgreement",
            "Precision",
            "Recall",
            "F1 Score",
        ][:len(metrics_display.columns)]

        metrics_table = metrics_axis.table(
            cellText=metrics_display.fillna("").values,
            colLabels=metrics_labels,
            cellLoc="center",
            loc="center",
            bbox=[0.01, 0.08, 0.98, 0.78],
        )

        metrics_table.auto_set_font_size(False)
        metrics_table.set_fontsize(9)

        style_header(
            metrics_table,
            len(metrics_display.columns),
        )

        count_columns = [
            "Approach",
            "True Positive (TP)",
            "False Positive (FP)",
            "False Negative (FN)",
            "True Negative (TN)",
            "Total Fields",
            "Matched Fields",
            "Incorrect Values",
            "Missing Values",
            "Unexpected Values",
        ]

        counts_display = summary_df[
            [
                column
                for column in count_columns
                if column in summary_df.columns
            ]
        ].copy()

        count_labels = {
            "Approach": "Approach",
            "True Positive (TP)": "True Positive\n(TP)",
            "False Positive (FP)": "False Positive\n(FP)",
            "False Negative (FN)": "False Negative\n(FN)",
            "True Negative (TN)": "True Negative\n(TN)",
            "Total Fields": "Total\nFields",
            "Matched Fields": "Exact\nMatches",
            "Incorrect Values": "Incorrect\nValues",
            "Missing Values": "Missing\nValues",
            "Unexpected Values": "Unexpected\nValues",
        }

        counts_table = outcomes_axis.table(
            cellText=counts_display.fillna("").values,
            colLabels=[
                count_labels.get(
                    column,
                    column,
                )
                for column in counts_display.columns
            ],
            cellLoc="center",
            loc="center",
            bbox=[0.01, 0.08, 0.98, 0.78],
        )

        counts_table.auto_set_font_size(False)
        counts_table.set_fontsize(8)

        style_header(
            counts_table,
            len(counts_display.columns),
        )

        pdf.savefig(
            figure,
        )

        plt.close(figure)

    def add_field_page(
        pdf,
        page_dataframe,
        page_number,
        total_pages,
        row_offset,
    ):
        figure, axis = plt.subplots(
            figsize=(16.5, 11.7)
        )

        figure.subplots_adjust(
            left=0.025,
            right=0.975,
            top=0.90,
            bottom=0.055,
        )

        axis.axis("off")

        figure.suptitle(
            (
                "Field-Level Report "
                f"({page_number} of {total_pages})"
            ),
            fontsize=18,
            fontweight="bold",
            y=0.965,
        )

        figure.text(
            0.5,
            0.925,
            (
                "Green = exact match | "
                "Red = incorrect, missing, or unexpected | "
                "Gray = Expected | Amber = not run"
            ),
            ha="center",
            fontsize=9,
        )

        visible_dataframe = page_dataframe[
            visible_field_columns
        ].copy()

        for column in visible_dataframe.columns:
            wrap_width = (
                28
                if column == "Field"
                else 22
            )

            visible_dataframe[column] = (
                visible_dataframe[column]
                .apply(
                    lambda value: display_text(
                        value,
                        wrap_width,
                    )
                )
            )

        column_labels = [
            "Field",
            "Expected",
            "Plain\nPython",
            "OCR +\nPython",
            "Custom ML +\nPython",
            "LLM Schema\nMapping",
        ]

        table = axis.table(
            cellText=visible_dataframe.values,
            colLabels=column_labels,
            cellLoc="left",
            loc="center",
            bbox=[0.005, 0.03, 0.99, 0.85],
        )

        table.auto_set_font_size(False)
        table.set_fontsize(7.5)

        style_header(
            table,
            len(column_labels),
        )

        column_widths = [
            0.20,
            0.16,
            0.16,
            0.16,
            0.16,
            0.16,
        ]

        for (
            row_index,
            column_index,
        ), cell in table.get_celld().items():
            cell.set_width(
                column_widths[column_index]
            )

            cell.get_text().set_wrap(
                True
            )

            cell.get_text().set_verticalalignment(
                "center"
            )

            if row_index > 0:
                cell.set_height(
                    0.095
                )

        for local_row_index in range(
            len(visible_dataframe)
        ):
            table_row = local_row_index + 1

            source_row = field_df.iloc[
                row_offset + local_row_index
            ]

            field_cell = table[
                table_row,
                0,
            ]

            field_cell.set_facecolor(
                colors["field"]
            )

            field_cell.get_text().set_weight(
                "bold"
            )

            expected_cell = table[
                table_row,
                1,
            ]

            expected_cell.set_facecolor(
                colors["expected"]
            )

            expected_cell.get_text().set_weight(
                "bold"
            )

            for approach_column, approach in (
                enumerate(
                    approach_columns,
                    start=2,
                )
            ):
                status = str(
                    source_row.get(
                        status_columns[approach],
                        "Not Run",
                    )
                )

                target_cell = table[
                    table_row,
                    approach_column,
                ]

                if status in {
                    "Correct",
                    "Correct Blank",
                }:
                    target_cell.set_facecolor(
                        colors["correct"]
                    )

                elif status in {
                    "Incorrect",
                    "Missing",
                    "Unexpected",
                }:
                    target_cell.set_facecolor(
                        colors["incorrect"]
                    )

                else:
                    target_cell.set_facecolor(
                        colors["pending"]
                    )

        pdf.savefig(
            figure,
        )

        plt.close(figure)

    with PdfPages(output) as pdf:
        add_summary_page(pdf)

        rows_per_page = 8

        total_pages = max(
            1,
            (
                len(field_df)
                + rows_per_page
                - 1
            )
            // rows_per_page,
        )

        for page_index in range(
            total_pages
        ):
            start_row = (
                page_index * rows_per_page
            )

            end_row = min(
                start_row + rows_per_page,
                len(field_df),
            )

            add_field_page(
                pdf,
                field_df.iloc[
                    start_row:end_row
                ],
                page_index + 1,
                total_pages,
                start_row,
            )

    output.seek(0)
    return output



compare_tab, upload_tab, extract_tab = st.tabs(
    [
        "1. Compare & Export",
        "2. Upload PDF & Expected Values",
        "3. Individual Extraction (Optional)",
    ]
)


with upload_tab:
    st.header("Upload PDF and Expected Values")

    st.info(
        "Expected values are used only for evaluation. "
        "They are never supplied to Python, OCR, "
        "Custom ML, or LLM extraction."
    )

    (
        pdf_upload_tab,
        expected_json_tab,
        expected_manual_tab,
    ) = st.tabs(
        [
            "Upload Filled PDF",
            "Upload Expected JSON",
            "Enter Expected Values in CIOMS Form",
        ]
    )

    with pdf_upload_tab:
        st.subheader("Upload Filled CIOMS PDF")

        uploaded_pdf = st.file_uploader(
            "Upload filled CIOMS PDF",
            type=["pdf"],
            key="filled_pdf_upload",
        )

        if uploaded_pdf is not None:
            uploaded_bytes = uploaded_pdf.getvalue()
            uploaded_hash = hashlib.md5(
                uploaded_bytes
            ).hexdigest()

            if (
                st.session_state.get(
                    "uploaded_pdf_hash"
                )
                != uploaded_hash
            ):
                Path(
                    "data/filled_form.pdf"
                ).write_bytes(uploaded_bytes)

                clear_previous_outputs()

                st.session_state[
                    "uploaded_pdf_hash"
                ] = uploaded_hash

                st.session_state[
                    "expected_values"
                ] = blank_expected_values()

                st.session_state[
                    "expected_source"
                ] = "Not provided"

                st.success(
                    "PDF uploaded successfully. "
                    "Previous extraction outputs and "
                    "Expected values were cleared."
                )
            else:
                st.info(
                    "The same uploaded PDF is already "
                    "loaded."
                )

        current_session_pdf_ready = (
            bool(
                st.session_state.get(
                    "uploaded_pdf_hash"
                )
            )
            and Path(
                "data/filled_form.pdf"
            ).is_file()
        )

        if current_session_pdf_ready:
            st.success(
                "Filled CIOMS PDF is ready for "
                "extraction."
            )

            with open(
                "data/filled_form.pdf",
                "rb",
            ) as file:
                st.download_button(
                    "Download Uploaded PDF",
                    file,
                    file_name=(
                        "uploaded_filled_form.pdf"
                    ),
                    mime="application/pdf",
                )


        else:
            st.warning(
                "No filled CIOMS PDF has been "
                "uploaded."
            )

    with expected_json_tab:
        st.subheader(
            "Upload Verified Expected JSON"
        )

        st.caption(
            "The JSON must contain exactly the same "
            "38 keys defined by the CIOMS schema. "
            "Checkbox values must be Yes or Off."
        )

        template_json = json.dumps(
            blank_expected_values(),
            indent=2,
        )

        st.download_button(
            "Download Expected JSON Template",
            data=template_json,
            file_name=(
                "cioms_expected_values_template.json"
            ),
            mime="application/json",
        )

        uploaded_expected_json = st.file_uploader(
            "Upload expected-values JSON",
            type=["json"],
            key="expected_json_upload",
        )

        if uploaded_expected_json is not None:
            try:
                uploaded_expected = json.loads(
                    uploaded_expected_json
                    .getvalue()
                    .decode("utf-8-sig")
                )

                if not isinstance(
                    uploaded_expected,
                    dict,
                ):
                    raise ValueError(
                        "The JSON root must be an "
                        "object."
                    )

                validation = (
                    validate_expected_values(
                        uploaded_expected
                    )
                )

                has_errors = any(
                    validation.values()
                )

                if has_errors:
                    if validation[
                        "missing_keys"
                    ]:
                        st.error(
                            "Missing keys: "
                            + ", ".join(
                                validation[
                                    "missing_keys"
                                ]
                            )
                        )

                    if validation[
                        "unexpected_keys"
                    ]:
                        st.error(
                            "Unexpected keys: "
                            + ", ".join(
                                validation[
                                    "unexpected_keys"
                                ]
                            )
                        )

                    if validation[
                        "invalid_checkboxes"
                    ]:
                        st.error(
                            "Checkbox values must be "
                            "Yes or Off: "
                            + ", ".join(
                                validation[
                                    "invalid_checkboxes"
                                ]
                            )
                        )
                else:
                    expected_from_json = (
                        normalize_expected_values(
                            uploaded_expected
                        )
                    )

                    if st.button(
                        "Use Uploaded Expected JSON",
                        key=(
                            "use_uploaded_expected_json"
                        ),
                        type="primary",
                    ):
                        st.session_state[
                            "expected_values"
                        ] = expected_from_json

                        st.session_state[
                            "expected_source"
                        ] = (
                            "Verified JSON Upload"
                        )

                        st.success(
                            "Expected values loaded "
                            "from the verified JSON."
                        )

            except (
                UnicodeDecodeError,
                json.JSONDecodeError,
                ValueError,
            ) as error:
                st.error(
                    "Expected JSON is invalid: "
                    f"{error}"
                )

    with expected_manual_tab:
        st.subheader(
            "Manual Expected Values: CIOMS Form"
        )

        st.caption(
            "Enter values directly in their original "
            "CIOMS positions. Checkbox groups follow "
            "the form layout. The values are retained "
            "only in the current Streamlit session."
        )

        manual_values = (
            render_cioms_expected_form(
                initial_values=st.session_state[
                    "expected_values"
                ],
                key="manual_cioms_expected_form",
            )
        )

        manual_left, manual_middle, manual_right = (
            st.columns([1, 1, 1])
        )

        with manual_left:
            if st.button(
                "Save Manual Expected Values",
                key="save_manual_expected_values",
                type="primary",
                use_container_width=True,
            ):
                st.session_state[
                    "expected_values"
                ] = normalize_expected_values(
                    manual_values
                )

                st.session_state[
                    "expected_source"
                ] = (
                    "Manual CIOMS Form Entry"
                )

                st.success(
                    "Manual Expected values saved."
                )

        with manual_middle:
            if st.button(
                "Reset Manual Expected Values",
                key="reset_manual_expected_values",
                use_container_width=True,
            ):
                st.session_state[
                    "expected_values"
                ] = blank_expected_values()

                st.session_state[
                    "expected_source"
                ] = "Not provided"

                st.rerun()

        with manual_right:
            manual_download = json.dumps(
                normalize_expected_values(
                    manual_values
                ),
                indent=2,
            )

            st.download_button(
                "Download Manual Values as JSON",
                data=manual_download,
                file_name=(
                    "cioms_manual_expected_values.json"
                ),
                mime="application/json",
                use_container_width=True,
            )

    st.divider()

    st.write(
        "**Expected Values Source:** "
        f"{st.session_state['expected_source']}"
    )

    normalized_expected = (
        normalize_expected_values(
            st.session_state[
                "expected_values"
            ]
        )
    )

    populated_text_count = sum(
        bool(str(normalized_expected[field]).strip())
        for field in ALL_FIELDS
        if field not in CHECKBOX_SCHEMA
    )

    checked_count = sum(
        normalized_expected[field] == "Yes"
        for field in CHECKBOX_SCHEMA
    )

    st.caption(
        f"{populated_text_count} text fields populated; "
        f"{checked_count} checkboxes selected."
    )

    with st.expander(
        "Preview Active Expected Values"
    ):
        st.json(normalized_expected)


with extract_tab:
    st.header("Individual Extraction (Optional)")

    plain_tab, ocr_tab, ml_tab, llm_tab = st.tabs(
        [
            "Python",
            "OCR + Python",
            "Custom ML + Python",
            "LLM + Python",
        ]
    )

    with plain_tab:
        st.subheader("Plain Python Extraction")

        st.info(
            "Reads embedded PDF form-widget names "
            "and values using PyMuPDF, then maps "
            "them into the common 38-field CIOMS "
            "schema. This approach requires a "
            "fillable PDF with accessible widgets."
        )

        if st.button("Run Python Extraction"):
            if not Path("data/filled_form.pdf").exists():
                st.error("Please upload the filled PDF form first.")
            else:
                result = run_with_progress(
                    "Python Extraction",
                    run_plain_python
                )
                st.json(clean_display_output(result))

    with ocr_tab:
        st.subheader("OCR + Python Extraction")

        st.info(
            "Renders the PDF page visually, applies "
            "OCR to recognize visible content, and "
            "uses Python logic to map the recognized "
            "values into the 38-field CIOMS schema."
        )

        if st.button("Run OCR Extraction"):
            if not Path("data/filled_form.pdf").exists():
                st.error("Please upload the filled PDF form first.")
            else:
                result = run_with_progress(
                    "OCR extraction",
                    run_ocr_python
                )

                if str(result.get("ocr_status", "")).lower().startswith("failed"):
                    st.warning(result.get("ocr_status"))

                st.json(clean_display_output(result))

        if Path("outputs/reports/ocr_raw_text.txt").exists():
            with st.expander("View OCR Raw Text"):
                st.text(
                    Path("outputs/reports/ocr_raw_text.txt").read_text(
                        encoding="utf-8"
                    )
                )

    with ml_tab:
        st.subheader("Custom ML + Python Extraction")

        st.info(
            "Processes rendered PDF pixels using a "
            "fine-tuned CRNN for 23 text fields and "
            "a separately trained checkbox CNN for "
            "15 checkbox fields. The models do not "
            "receive Expected Values during inference."
        )

        if st.button("Run Custom ML Extraction"):
            if not Path("data/filled_form.pdf").exists():
                st.error("Please upload the filled PDF form first.")
            else:
                result = run_with_progress(
                    "Custom ML extraction",
                    run_custom_ml
                )
                st.json(clean_display_output(result))

    with llm_tab:
        st.subheader("LLM Schema Mapping")

        st.info(
            "Uses an LLM to map Python-extracted PDF "
            "form-field values into the required "
            "38-field JSON schema. This is schema "
            "mapping, not independent visual PDF "
            "extraction."
        )

        if st.button("Run LLM Extraction"):
            if not Path("data/filled_form.pdf").exists():
                st.error("Please upload the filled PDF form first.")
            else:
                result = run_with_progress(
                    "LLM extraction",
                    run_llm
                )

                if result.get("llm_status", "").lower() != "success":
                    st.warning(
                        result.get(
                            "llm_status",
                            "LLM extraction completed with warning."
                        )
                    )

                st.json(clean_display_output(result))


with compare_tab:
    st.header("Compare & Export Results")

    st.write(
        "All four approaches are evaluated against "
        "the independently provided Expected Values "
        "using normalized exact-match field-level "
        "evaluation."
    )

    st.subheader("Extraction Approach Briefs")

    (
        plain_brief_column,
        ocr_brief_column,
        ml_brief_column,
        llm_brief_column,
    ) = st.columns(4)

    with plain_brief_column:
        st.markdown("#### Plain Python")
        st.write(
            "Reads embedded PDF form-widget names "
            "and values with PyMuPDF, then maps them "
            "to the common 38-field CIOMS schema."
        )

    with ocr_brief_column:
        st.markdown("#### OCR + Python")
        st.write(
            "Renders the PDF visually, performs OCR, "
            "and uses Python logic to map recognized "
            "content into the CIOMS field schema."
        )

    with ml_brief_column:
        st.markdown("#### Custom ML + Python")
        st.write(
            "Uses the fine-tuned CRNN for 23 text "
            "fields and the custom checkbox CNN for "
            "15 checkbox fields from rendered pixels."
        )

    with llm_brief_column:
        st.markdown("#### LLM Schema Mapping")
        st.write(
            "Uses an LLM to map Python-extracted PDF "
            "form-field values into the required "
            "38-field JSON schema."
        )

    st.caption(
        "Current Expected Values source: "
        f"{st.session_state['expected_source']}"
    )

    pdf_is_ready = (
        bool(
            st.session_state.get(
                "uploaded_pdf_hash"
            )
        )
        and Path(
            "data/filled_form.pdf"
        ).is_file()
    )

    expected_values_are_ready = (
        st.session_state["expected_source"]
        != "Not provided"
    )

    comparison_is_ready = (
        pdf_is_ready
        and expected_values_are_ready
    )

    st.subheader("Comparison Readiness")

    readiness_left, readiness_right = (
        st.columns(2)
    )

    with readiness_left:
        if pdf_is_ready:
            st.success(
                "Filled CIOMS PDF uploaded"
            )
        else:
            st.error(
                "Filled CIOMS PDF not uploaded"
            )

    with readiness_right:
        if expected_values_are_ready:
            st.success(
                "Expected Values provided"
            )
        else:
            st.error(
                "Expected Values not provided"
            )

    if not comparison_is_ready:
        st.info(
            "Complete both readiness requirements "
            "to enable Extract All Approaches & "
            "Run Comparison."
        )

    if st.button(
        "Extract All Approaches & Run Comparison",
        type="primary",
        disabled=not comparison_is_ready,
        help=(
            None
            if comparison_is_ready
            else (
                "Upload a filled CIOMS PDF and "
                "provide Expected Values first."
            )
        ),
    ):
        if not Path(
            "data/filled_form.pdf"
        ).exists():
            st.error(
                "Please upload the filled CIOMS PDF "
                "first."
            )

        elif (
            st.session_state[
                "expected_source"
            ]
            == "Not provided"
        ):
            st.error(
                "Please upload a verified Expected "
                "JSON or save values from the manual "
                "CIOMS form before comparison."
            )

        else:
            active_expected = (
                normalize_expected_values(
                    st.session_state[
                        "expected_values"
                    ]
                )
            )

            save_json(
                active_expected,
                "outputs/json/expected_values.json",
            )

            try:
                run_with_progress(
                    "Plain Python extraction",
                    run_plain_python,
                )

                run_with_progress(
                    "OCR + Python extraction",
                    run_ocr_python,
                )

                run_with_progress(
                    "Custom ML + Python extraction",
                    run_custom_ml,
                )

                run_with_progress(
                    "LLM schema mapping",
                    run_llm,
                )

                run_with_progress(
                    "Expected-value comparison",
                    lambda: compare_all_outputs(
                        expected_values_path=(
                            "outputs/json/"
                            "expected_values.json"
                        ),
                        expected_source=(
                            st.session_state[
                                "expected_source"
                            ]
                        ),
                    ),
                )

                field_df = (
                    build_field_level_dataframe()
                )

                confusion_df = (
                    build_confusion_dataframe()
                )

                field_df.to_csv(
                    "outputs/reports/"
                    "field_level_report.csv",
                    index=False,
                )

                confusion_df.to_csv(
                    "outputs/reports/"
                    "confusion_summary.csv",
                    index=False,
                )

                st.success(
                    "Comparison results are ready "
                    "below."
                )

            except Exception as error:
                st.error(
                    "Comparison failed: "
                    f"{error}"
                )

    if (
        st.session_state["expected_source"]
        != "Not provided"
        and Path(
            "outputs/json/expected_values.json"
        ).is_file()
        and Path(
            "outputs/reports/comparison_report.csv"
        ).is_file()
    ):
        summary_df = pd.read_csv(
            "outputs/reports/comparison_report.csv"
        )

        summary_df = summary_df.rename(
            columns={
                "TP": "True Positive (TP)",
                "FP": "False Positive (FP)",
                "FN": "False Negative (FN)",
                "TN": "True Negative (TN)",
            }
        )

        summary_df = summary_df.drop(
            columns=["Wrong/Missing Fields"],
            errors="ignore",
        )

        field_df = (
            build_field_level_dataframe()
        )

        confusion_df = (
            build_confusion_dataframe()
        )

        st.subheader(
            "Approach Comparison Summary"
        )

        st.caption(
            "True Positive means a correct non-empty "
            "value. False Positive means an unexpected "
            "or incorrect produced value. False Negative "
            "means an expected value was not correctly "
            "recovered. True Negative means both "
            "Expected and extracted values are blank."
        )

        st.dataframe(
            summary_df,
            width="stretch",
            hide_index=True,
        )

        with st.expander(
            "Field-Level Report",
            expanded=True,
        ):
            st.caption(
                "Green indicates an exact normalized "
                "match. Red indicates an incorrect, "
                "missing, or unexpected value. Gray "
                "contains the independently supplied "
                "Expected value."
            )

            st.dataframe(
                style_field_level_report(
                    field_df
                ),
                width="stretch",
                hide_index=True,
            )

        st.subheader("Export Results")

        csv_bytes = field_df.to_csv(index=False).encode("utf-8-sig")
        excel_bytes = create_excel_bytes(summary_df, field_df, confusion_df)
        pdf_bytes = create_pdf_bytes(summary_df, field_df, confusion_df)

        col1, col2, col3 = st.columns(3)

        with col1:
            st.download_button(
                "Download CSV",
                csv_bytes,
                file_name="field_level_comparison_report.csv",
                mime="text/csv"
            )

        with col2:
            st.download_button(
                "Download Excel",
                excel_bytes,
                file_name="extraction_quality_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col3:
            st.download_button(
                "Download PDF",
                pdf_bytes,
                file_name="extraction_quality_report.pdf",
                mime="application/pdf"
            )
